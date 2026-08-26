import logging
import os
import pandas as pd
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName


COR_PRIMARIA = "006B64"
COR_SECUNDARIA = "00A859"
COR_DESTAQUE = "D9EAD3"
COR_FUNDO = "F4F7F6"
COR_BORDA = "D9E2E1"

logger = logging.getLogger(__name__)


def _ultima_linha(dataframes, nome_aba):

    return len(dataframes[nome_aba]) + 1


def _formula_sumifs(
    aba,
    coluna_valor,
    ultima_linha,
    criterios
):

    argumentos = [
        f"'{aba}'!${coluna_valor}$2:${coluna_valor}${ultima_linha}"
    ]

    for coluna, criterio in criterios:
        argumentos.extend([
            f"'{aba}'!${coluna}$2:${coluna}${ultima_linha}",
            criterio
        ])

    return f"=SUMIFS({','.join(argumentos)})"


def _estilizar_tabela(worksheet, intervalo):

    borda = Side(style="thin", color=COR_BORDA)

    for cell in worksheet[intervalo][0]:
        cell.fill = PatternFill("solid", fgColor=COR_PRIMARIA)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    for linha in worksheet[intervalo]:
        for cell in linha:
            cell.border = Border(bottom=borda)


def _aplicar_formato_numero(worksheet, intervalo, formato):

    for linha in worksheet[intervalo]:
        for cell in linha:
            cell.number_format = formato


def _adicionar_grafico(
    worksheet,
    titulo,
    categoria_coluna,
    dados_coluna_inicial,
    dados_coluna_final,
    linha_cabecalho,
    linha_final,
    posicao,
    horizontal=False,
    largura=13,
    altura=7
):

    chart = BarChart()
    chart.type = "bar" if horizontal else "col"
    chart.style = 10
    chart.title = titulo
    chart.y_axis.title = "Valor" if not horizontal else "Categoria"
    chart.x_axis.title = "Categoria" if not horizontal else "Valor"
    chart.height = altura
    chart.width = largura

    dados = Reference(
        worksheet,
        min_col=dados_coluna_inicial,
        max_col=dados_coluna_final,
        min_row=linha_cabecalho,
        max_row=linha_final
    )
    categorias = Reference(
        worksheet,
        min_col=categoria_coluna,
        min_row=linha_cabecalho + 1,
        max_row=linha_final
    )

    chart.add_data(dados, titles_from_data=True)
    if dados_coluna_inicial == dados_coluna_final:
        chart.legend = None
    else:
        chart.legend.position = "b"
        for indice, coluna in enumerate(
            range(dados_coluna_inicial, dados_coluna_final + 1)
        ):
            chart.series[indice].tx = SeriesLabel(
                v=str(worksheet.cell(linha_cabecalho, coluna).value)
            )
    chart.set_categories(categorias)
    worksheet.add_chart(chart, posicao)


def _criar_resumo(writer, dataframes):

    logger.debug("Criação da aba de resumo iniciada")

    workbook = writer.book
    worksheet = workbook.create_sheet("resumo", 0)
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A8"
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.sheet_view.zoomScale = 85

    ultima_cadastro = _ultima_linha(dataframes, "dados_cadastrais")
    ultima_eventos = _ultima_linha(dataframes, "eventos_despesas")
    ultima_internacoes = _ultima_linha(dataframes, "internacoes")
    ultima_despesas = _ultima_linha(
        dataframes,
        "despesas_assistenciais"
    )
    ultima_graficos = _ultima_linha(
        dataframes,
        "indicadores_graficos"
    )

    worksheet.merge_cells("A1:T1")
    worksheet["A1"] = "Resumo dos Indicadores Assistenciais"
    worksheet["A1"].fill = PatternFill("solid", fgColor=COR_PRIMARIA)
    worksheet["A1"].font = Font(bold=True, color="FFFFFF", size=16)
    worksheet["A1"].alignment = Alignment(
        horizontal="left",
        vertical="center"
    )
    worksheet.row_dimensions[1].height = 30

    worksheet["A3"] = "Unimed selecionada"
    worksheet["A4"] = "Código Unimed"
    worksheet["A5"] = "Período"

    primeira_unimed = dataframes["dados_cadastrais"].iloc[0]
    worksheet["B3"] = primeira_unimed["unimed"]
    worksheet["B4"] = (
        f"=INDEX('dados_cadastrais'!$C$2:$C${ultima_cadastro},"
        f"MATCH(B3,'dados_cadastrais'!$A$2:$A${ultima_cadastro},0))"
    )
    worksheet["B5"] = (
        f"=INDEX('eventos_despesas'!$B$2:$B${ultima_eventos},"
        f"MATCH(B4,'eventos_despesas'!$A$2:$A${ultima_eventos},0))"
    )

    nome_lista = DefinedName(
        "lista_unimeds",
        attr_text=(
            f"'dados_cadastrais'!$A$2:$A${ultima_cadastro}"
        )
    )
    workbook.defined_names.add(nome_lista)

    validacao = DataValidation(
        type="list",
        formula1="=lista_unimeds",
        allow_blank=False
    )
    validacao.promptTitle = "Selecione uma Unimed"
    validacao.prompt = "Escolha uma Unimed cadastrada no relatório."
    validacao.error = "Selecione uma Unimed da lista."
    validacao.errorTitle = "Unimed inválida"
    validacao.showErrorMessage = True
    worksheet.add_data_validation(validacao)
    validacao.add(worksheet["B3"])

    for linha in worksheet["A3:A5"]:
        linha[0].font = Font(bold=True, color=COR_PRIMARIA)

    worksheet["B3"].fill = PatternFill("solid", fgColor="FFF2CC")
    worksheet["B3"].font = Font(bold=True)

    kpis = [
        (
            "D3", "Beneficiários", "E3",
            f"=INDEX('dados_cadastrais'!$I$2:$I${ultima_cadastro},"
            f"MATCH($B$3,'dados_cadastrais'!$A$2:$A${ultima_cadastro},0))"
        ),
        (
            "F3", "Total de Eventos", "G3",
            _formula_sumifs(
                "eventos_despesas", "E", ultima_eventos,
                [
                    ("A", "$B$4"),
                    ("B", "$B$5"),
                    ("C", '"Total Eventos"'),
                    ("D", '"EVENTOS"')
                ]
            )
        ),
        (
            "H3", "Despesas SIP", "I3",
            _formula_sumifs(
                "eventos_despesas", "E", ultima_eventos,
                [
                    ("A", "$B$4"),
                    ("B", "$B$5"),
                    ("C", '"Total Despesas Assistenciais R$"'),
                    ("D", '"DESPESA"')
                ]
            )
        ),
        (
            "J3", "Internações", "K3",
            _formula_sumifs(
                "internacoes", "D", ultima_internacoes,
                [("A", "$B$4"), ("B", "$B$5")]
            )
        ),
        (
            "L3", "Despesas DIOPS", "M3",
            _formula_sumifs(
                "despesas_assistenciais", "D", ultima_despesas,
                [("A", "$B$4"), ("B", "$B$5")]
            )
        )
    ]

    for celula_rotulo, rotulo, celula_valor, formula in kpis:
        worksheet[celula_rotulo] = rotulo
        worksheet[celula_valor] = formula
        worksheet[celula_rotulo].fill = PatternFill(
            "solid",
            fgColor=COR_PRIMARIA
        )
        worksheet[celula_rotulo].font = Font(
            bold=True,
            color="FFFFFF"
        )
        worksheet[celula_valor].fill = PatternFill(
            "solid",
            fgColor=COR_DESTAQUE
        )
        worksheet[celula_valor].font = Font(
            bold=True,
            color=COR_PRIMARIA,
            size=12
        )
        worksheet[celula_valor].number_format = "#,##0"

    categorias_eventos = [
        "Consultas (Amb+PS)",
        "Outros Atend. Amb.",
        "Exames",
        "Terapias",
        "Internações",
        "Demais Desp Amb e Hosp"
    ]

    worksheet["A8"] = "Categoria"
    worksheet["B8"] = "Eventos"
    worksheet["C8"] = "Despesas SIP (R$)"

    for linha, categoria in enumerate(categorias_eventos, start=9):
        worksheet.cell(linha, 1, categoria)
        worksheet.cell(
            linha,
            2,
            _formula_sumifs(
                "eventos_despesas", "E", ultima_eventos,
                [
                    ("A", "$B$4"),
                    ("B", "$B$5"),
                    ("C", f"$A{linha}"),
                    ("D", '"EVENTOS"')
                ]
            )
        )
        worksheet.cell(
            linha,
            3,
            _formula_sumifs(
                "eventos_despesas", "E", ultima_eventos,
                [
                    ("A", "$B$4"),
                    ("B", "$B$5"),
                    ("C", f"$A{linha}"),
                    ("D", '"DESPESA"')
                ]
            )
        )

    _estilizar_tabela(worksheet, "A8:C14")
    _aplicar_formato_numero(worksheet, "B9:C14", "#,##0")

    tipos_internacao = [
        "Internações Clínicas",
        "Internações Cirúrgicas",
        "Obstétrica",
        "Internações Pediátricas",
        "Psiquiatria"
    ]

    worksheet["A18"] = "Tipo de internação"
    worksheet["B18"] = "Quantidade"

    for linha, tipo in enumerate(tipos_internacao, start=19):
        worksheet.cell(linha, 1, tipo)
        worksheet.cell(
            linha,
            2,
            _formula_sumifs(
                "internacoes", "D", ultima_internacoes,
                [
                    ("A", "$B$4"),
                    ("B", "$B$5"),
                    ("C", f"$A{linha}")
                ]
            )
        )

    _estilizar_tabela(worksheet, "A18:B23")
    _aplicar_formato_numero(worksheet, "B19:B23", "#,##0")

    tipos_despesa = list(
        dataframes["despesas_assistenciais"][
            "cd_tipo_despesa"
        ].drop_duplicates()
    )

    worksheet["A27"] = "Tipo de despesa assistencial"
    worksheet["B27"] = "Valor DIOPS (R$)"

    for linha, tipo in enumerate(tipos_despesa, start=28):
        worksheet.cell(linha, 1, tipo)
        worksheet.cell(
            linha,
            2,
            _formula_sumifs(
                "despesas_assistenciais", "D", ultima_despesas,
                [
                    ("A", "$B$4"),
                    ("B", "$B$5"),
                    ("C", f"$A{linha}")
                ]
            )
        )

    ultima_linha_despesas = 27 + len(tipos_despesa)
    _estilizar_tabela(
        worksheet,
        f"A27:B{ultima_linha_despesas}"
    )
    _aplicar_formato_numero(
        worksheet,
        f"B28:B{ultima_linha_despesas}",
        "#,##0"
    )

    tabelas_graficos = [
        (
            40,
            "Consultas por 1.000 beneficiários",
            "CONSULTAS_POR_MIL_BENEFICIARIOS",
            [
                ("CONSULTAS_MEDICAS_AMB_PS", "Consultas médicas"),
                ("CONSULTAS_AMBULATORIAIS", "Consultas ambulatoriais"),
                ("CONSULTAS_PRONTO_SOCORRO", "Pronto-socorro")
            ],
            "G40"
        ),
        (
            57,
            "Exames por 1.000 beneficiários",
            "EXAMES_POR_MIL_BENEFICIARIOS",
            [
                ("RESSONANCIA_MAGNETICA", "Ressonância magnética"),
                (
                    "TOMOGRAFIA_COMPUTADORIZADA",
                    "Tomografia computadorizada"
                ),
                ("HEMOGLOBINA_GLICADA", "Hemoglobina glicada")
            ],
            "G57"
        ),
        (
            74,
            "Terapias e outros atendimentos por 1.000 beneficiários",
            (
                "TERAPIAS_OUTROS_ATENDIMENTOS_"
                "POR_MIL_BENEFICIARIOS"
            ),
            [
                ("TERAPIAS", "Terapias"),
                ("OUTROS_ATENDIMENTOS", "Outros atendimentos")
            ],
            "G74"
        )
    ]

    entidades = [
        ("UNIMED", "Unimed"),
        ("PORTE_REGIAO", "Porte - Região"),
        ("PORTE_NACIONAL", "Porte - Nacional"),
        ("MEDIA_NACIONAL_GERAL", "Média nacional")
    ]

    for (
        linha_inicial,
        titulo,
        cd_grafico,
        indicadores,
        posicao_grafico
    ) in tabelas_graficos:

        worksheet.cell(linha_inicial, 1, "Indicador")

        for coluna, (_, descricao) in enumerate(entidades, start=2):
            worksheet.cell(linha_inicial, coluna, descricao)

        for deslocamento, (cd_indicador, descricao) in enumerate(
            indicadores,
            start=1
        ):
            linha = linha_inicial + deslocamento
            worksheet.cell(linha, 1, descricao)

            for coluna, (tp_entidade, _) in enumerate(
                entidades,
                start=2
            ):
                worksheet.cell(
                    linha,
                    coluna,
                    _formula_sumifs(
                        "indicadores_graficos",
                        "G",
                        ultima_graficos,
                        [
                            ("A", "$B$4"),
                            ("B", "$B$5"),
                            ("C", f'"{cd_grafico}"'),
                            ("D", f'"{cd_indicador}"'),
                            ("E", f'"{tp_entidade}"')
                        ]
                    )
                )

        linha_final = linha_inicial + len(indicadores)
        _estilizar_tabela(
            worksheet,
            f"A{linha_inicial}:E{linha_final}"
        )
        _aplicar_formato_numero(
            worksheet,
            f"B{linha_inicial + 1}:E{linha_final}",
            "#,##0"
        )

        _adicionar_grafico(
            worksheet=worksheet,
            titulo=titulo,
            categoria_coluna=1,
            dados_coluna_inicial=2,
            dados_coluna_final=5,
            linha_cabecalho=linha_inicial,
            linha_final=linha_final,
            posicao=posicao_grafico,
            largura=18,
            altura=8
        )

    worksheet["A91"] = "Indicador DIOPS"
    worksheet["B91"] = "Valor (R$)"

    indicadores_tiquete_custo = [
        ("TIQUETE_MEDIO", "TM - Tíquete Médio"),
        ("CUSTO_PER_CAPITA", "CP - Custo Per Capita")
    ]

    for linha, (cd_indicador, descricao) in enumerate(
        indicadores_tiquete_custo,
        start=92
    ):
        worksheet.cell(linha, 1, descricao)
        worksheet.cell(
            linha,
            2,
            _formula_sumifs(
                "indicadores_graficos",
                "G",
                ultima_graficos,
                [
                    ("A", "$B$4"),
                    ("B", "$B$5"),
                    (
                        "C",
                        '"TIQUETE_MEDIO_CUSTO_PER_CAPITA_DIOPS"'
                    ),
                    ("D", f'"{cd_indicador}"'),
                    ("E", '"UNIMED"')
                ]
            )
        )

    _estilizar_tabela(worksheet, "A91:B93")
    _aplicar_formato_numero(worksheet, "B92:B93", "R$ #,##0.00")

    _adicionar_grafico(
        worksheet,
        "Tíquete Médio e Custo Per Capita (Base DIOPS)",
        1, 2, 2, 91, 93, "G91"
    )

    _adicionar_grafico(
        worksheet,
        "Eventos por categoria",
        1, 2, 2, 8, 14, "E8"
    )
    _adicionar_grafico(
        worksheet,
        "Despesas SIP por categoria (R$)",
        1, 3, 3, 8, 14, "M8"
    )
    _adicionar_grafico(
        worksheet,
        "Internações por tipo",
        1, 2, 2, 18, 23, "E24"
    )
    _adicionar_grafico(
        worksheet,
        "Despesas DIOPS por tipo (R$)",
        1, 2, 2, 27, ultima_linha_despesas, "M24",
        horizontal=True
    )

    worksheet.column_dimensions["A"].width = 58
    worksheet.column_dimensions["B"].width = 24
    worksheet.column_dimensions["C"].width = 20

    for coluna in range(4, 21):
        worksheet.column_dimensions[get_column_letter(coluna)].width = 14

    worksheet["A1"].fill = PatternFill("solid", fgColor=COR_PRIMARIA)
    worksheet["A1"].font = Font(
        name="Aptos Display",
        bold=True,
        color="FFFFFF",
        size=16
    )

    for intervalo in [
        "A8:C14",
        "A18:B23",
        f"A27:B{ultima_linha_despesas}",
        "A40:E43",
        "A57:E60",
        "A74:E76",
        "A91:B93"
    ]:
        _estilizar_tabela(worksheet, intervalo)

    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    logger.debug("Aba de resumo criada | graficos=8")

def export_to_excel(dados, output_path):

    logger.info(
        "Geração do Excel iniciada | arquivo_saida=%s | abas_dados=%d",
        output_path,
        len(dados)
    )

    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    dataframes = {
        nome_aba: pd.DataFrame(registros)
        for nome_aba, registros in dados.items()
    }

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

        for nome_aba, dataframe in dataframes.items():
            logger.debug(
                "Escrevendo aba | nome=%s | registros=%d",
                nome_aba,
                len(dataframe)
            )
            dataframe.to_excel(
                writer,
                sheet_name=nome_aba,
                index=False
            )

            worksheet = writer.sheets[nome_aba]
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
            worksheet.sheet_view.showGridLines = False

            for cell in worksheet[1]:
                cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor="006B64"
                )
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

            for indice_coluna, coluna in enumerate(
                worksheet.iter_cols(),
                start=1
            ):
                largura = min(
                    max(
                        len(str(celula.value))
                        if celula.value is not None
                        else 0
                        for celula in coluna
                    ) + 2,
                    45
                )

                worksheet.column_dimensions[
                    get_column_letter(indice_coluna)
                ].width = largura

        _criar_resumo(writer, dataframes)

    logger.info(
        "Geração do Excel concluída | arquivo_saida=%s | "
        "total_registros=%d",
        output_path,
        sum(len(dataframe) for dataframe in dataframes.values())
    )

    return dataframes
