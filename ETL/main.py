import os

# BIBLIOTECAS
import pandas as pd
from dotenv import load_dotenv
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# EXTRACTORS
from ETL.extractors.expenses import ExpensesExtractor
from extractors.events import EventsExtractor
from extractors.header import HeaderExtractor
from extractors.hospitalizations import HospitalizationsExtractor
from extractors.indicators import IndicatorsExtractor

# TRANSFORMERS
from transformers.events import EventsTransformer
from transformers.hospitalizations import HospitalizationsTransformer
from transformers.indicators import IndicatorsTransformer

from readers.pdf_reader import PDFReader


load_dotenv()

FILE_NAME = os.getenv("FILE_NAME_2025T4_RS") # Change this to the desired file name from .env

PDF_PATH = os.path.abspath(
    f"ETL/{os.getenv('FILE_PATH')}{FILE_NAME}.pdf"
)

OUTPUT_DIR = os.path.abspath("ETL/output")
OUTPUT_PATH = os.path.join(
    OUTPUT_DIR,
    f"{FILE_NAME}_extraido.xlsx"
)


def extract_all(reader, paginas):

    header_extractor = HeaderExtractor()
    events_extractor = EventsExtractor()
    expenses_extractor = ExpensesExtractor()
    indicators_extractor = IndicatorsExtractor()
    hospitalizations_extractor = HospitalizationsExtractor()

    events_transformer = EventsTransformer()
    # expenses_transformer = ExpensesTransformer()
    indicators_transformer = IndicatorsTransformer()
    hospitalizations_transformer = HospitalizationsTransformer()

    dados_cadastrais = []
    eventos_despesas = []
    indicadores_assistenciais = []
    internacoes = []

    for pagina in paginas:

        texto = reader.get_page_text(pagina)
        header = header_extractor.extract(texto)

        dados_eventos = events_extractor.extract(texto)
        dados_indicadores = indicators_extractor.extract(texto)

        texto_internacoes = reader.get_page_region_text(
            page_index=pagina + 2,
            start_text="Número de Internações",
            end_text="Total Internações"
        )

        dados_internacoes = hospitalizations_extractor.extract(
            texto_internacoes
        )

        dados_cadastrais.append(header)

        eventos_despesas.extend(
            events_transformer.transform(dados_eventos, header)
        )

        indicadores_assistenciais.extend(
            indicators_transformer.transform(dados_indicadores, header)
        )

        internacoes.extend(
            hospitalizations_transformer.transform(
                dados_internacoes,
                header
            )
        )

    return {
        "dados_cadastrais": dados_cadastrais,
        "eventos_despesas": eventos_despesas,
        "indicadores_assistenciais": indicadores_assistenciais,
        "internacoes": internacoes
    }


def export_to_excel(dados, output_path):

    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    dataframes = {
        nome_aba: pd.DataFrame(registros)
        for nome_aba, registros in dados.items()
    }

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

        for nome_aba, dataframe in dataframes.items():
            dataframe.to_excel(
                writer,
                sheet_name=nome_aba,
                index=False
            )

            worksheet = writer.sheets[nome_aba]
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
            worksheet.sheet_view.showGridLines = False

            for cell in worksheet[1]:
                cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor="006B64"
                )
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

            for indice_coluna, coluna in enumerate(
                worksheet.iter_cols(),
                start=1
            ):
                largura = min(
                    max(
                        len(str(celula.value))
                        if celula.value is not None
                        else 0
                        for celula in coluna
                    ) + 2,
                    45
                )

                worksheet.column_dimensions[
                    get_column_letter(indice_coluna)
                ].width = largura

    return dataframes


def main():

    reader = PDFReader(PDF_PATH)
    paginas = reader.locate_pages()

    dados = extract_all(reader, paginas)
    dataframes = export_to_excel(dados, OUTPUT_PATH)

    print()
    print("=" * 50)
    print("EXCEL GERADO")
    print("=" * 50)
    print(OUTPUT_PATH)

    for nome_aba, dataframe in dataframes.items():
        print(f"{nome_aba}: {len(dataframe)} registros")


if __name__ == "__main__":
    main()
